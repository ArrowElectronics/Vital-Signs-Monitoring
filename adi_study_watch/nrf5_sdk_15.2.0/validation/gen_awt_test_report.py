import os
import common
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill


def analyze_data(usecase_csv_dir, stream):
    """

    :param usecase_csv_dir:
    :param stream:
    :return:
    """
    results_dict = None
    for file_name in os.listdir(usecase_csv_dir):
        if stream == 'ppg':
            if 'adpdappstream' in file_name.lower():
                usecase_log_file_path = os.path.join(usecase_csv_dir, file_name)
                results_dict = common.analyze_wfm(usecase_log_file_path, 'ppg', 'awt', '0')
                break
        elif stream == 'ecg':
            if 'ecgappstream' in file_name.lower():
                usecase_log_file_path = os.path.join(usecase_csv_dir, file_name)
                results_dict = common.analyze_wfm(usecase_log_file_path, 'ecg', 'awt', '0')
                break
    return results_dict


def analyze_watch_awt_results(report_dir):
    """

    :param report_dir:
    :return:
    """
    watch_results_dict = {}
    for watch_addr in os.listdir(report_dir):
        results_dict = {'uc1_ppg':None, 'uc2_ppg':None, 'uc3_ppg':None, 'uc4_ppg':None, 'uc5_ppg':None,
                        'uc1_ecg':None, 'uc2_ecg':None, 'uc3_ecg':None, 'uc4_ecg':None, 'uc5_ecg':None}
        watch_dir = os.path.join(report_dir, watch_addr)
        for usecase in os.listdir(watch_dir):
            usecase_csv_dir = os.path.join(watch_dir, usecase, 'csv')
            if usecase.lower() == 'usecase1':  # PPG
                results_dict['uc1_ppg'] = analyze_data(usecase_csv_dir, 'ppg')
            elif usecase.lower() == 'usecase2':  # PPG
                results_dict['uc2_ppg'] = analyze_data(usecase_csv_dir, 'ppg')
            elif usecase.lower() == 'usecase3':  # PPG + ECG
                results_dict['uc3_ppg'] = analyze_data(usecase_csv_dir, 'ppg')
                results_dict['uc3_ecg'] = analyze_data(usecase_csv_dir, 'ecg')
            elif usecase.lower() == 'usecase4':  # PPG + ECG
                results_dict['uc4_ppg'] = analyze_data(usecase_csv_dir, 'ppg')
                results_dict['uc4_ecg'] = analyze_data(usecase_csv_dir, 'ecg')
            elif usecase.lower() == 'usecase5':  # PPG
                results_dict['uc5_ppg'] = analyze_data(usecase_csv_dir, 'ppg')
            else:
                pass
        watch_results_dict[watch_addr] = results_dict
    return watch_results_dict


def generate_awt_test_results(report_dir):
    """

    :param report_dir:
    :return:
    """
    watch_results_dict = analyze_watch_awt_results(report_dir)

    curr_dir = os.getcwd()
    template_path = os.path.join(curr_dir, 'awt_usecase_report_template.xlsx')
    report_path = os.path.join(report_dir, 'awt_test_report.xlsx')
    row_offset = 4
    wb = load_workbook(template_path)
    ws = wb.get_active_sheet()
    row_idx = row_offset
    for watch_addr, result_dict in watch_results_dict.items():
        col_idx = 1
        ws.cell(row=row_idx, column=col_idx).value = watch_addr
        col_idx += 1
        for i in range(5):
            ws.cell(row=row_idx, column=col_idx).value = 'Usecase{}'.format(i+1)
            # PPG Results
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i+1)]['mi_ch0']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['stdev_raw_ch0']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['dc_snr_db_ch0']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['dc_amp_ch0']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['ac_snr_db_ch0']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['ac_amp_ch0']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['peak_count_ch0']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['trough_count_ch0']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['mi_ch1']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['stdev_raw_ch1']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['dc_snr_db_ch1']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['dc_amp_ch1']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['ac_snr_db_ch1']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['ac_amp_ch1']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['peak_count_ch1']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ppg'.format(i + 1)]['trough_count_ch1']) if result_dict['uc{}_ppg'.format(i+1)] else ' - '
            # ECG results
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ecg'.format(i+1)]['hr_mean']) if result_dict['uc{}_ecg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ecg'.format(i+1)]['qrs_amp_mean']) if result_dict['uc{}_ecg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ecg'.format(i+1)]['qrs_time_mean']) if result_dict['uc{}_ecg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ecg'.format(i+1)]['qt_mean']) if result_dict['uc{}_ecg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ecg'.format(i+1)]['st_level_mean']) if result_dict['uc{}_ecg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ecg'.format(i+1)]['iso_mean']) if result_dict['uc{}_ecg'.format(i+1)] else ' - '
            col_idx += 1
            ws.cell(row=row_idx, column=col_idx).value = float(result_dict['uc{}_ecg'.format(i+1)]['pr_mean']) if result_dict['uc{}_ecg'.format(i+1)] else ' - '
            row_idx += 1
            col_idx = 2
        row_idx + 5
    wb.save(report_path)
    return report_path

if __name__ == '__main__':
    report_dir = common.easygui.diropenbox('', 'Select the Test Results folder path and press OK')
    if os.path.exists(report_dir):
        report_path = generate_awt_test_results(report_dir)
        common.easygui.msgbox('Report generation completed!\nReport Path: {}'.format(report_path))
    else:
        common.easygui.msgbox('Invalid folder path selected!')
