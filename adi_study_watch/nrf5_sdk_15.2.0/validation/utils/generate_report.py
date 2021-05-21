import os
import shutil
import common
import easygui
from utils import meas_check
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill

# ************** Global Variables ******************** #
shared_drive_dir = r'\\wilmnas4\Local Programs\FDWatch_TestData\Data_Testsetup\DVT1_Test_Results'
NA = ' - '
# **************************************************** #

def recur_list_dir(dir_path, empty_dir=True):
    for item in os.listdir(dir_path):
        item_path = os.path.join(dir_path, item)
        if os.path.isfile(item_path):
            empty_dir = False
        else:
            empty_dir = recur_list_dir(item_path, empty_dir)
    return empty_dir


def delete_empty_report_folders(dir_path):
    pcb_list = os.listdir(dir_path)
    for pcb in pcb_list:
        if '_test_result' in pcb:
            pcb_report_dir = os.path.join(shared_drive_dir, pcb)
            if os.path.exists(pcb_report_dir):
                for test_dir_name in os.listdir(pcb_report_dir):
                    test_dir = os.path.join(pcb_report_dir, test_dir_name)
                    if not os.path.isfile(test_dir):
                        empty_dir = recur_list_dir(test_dir)
                        if empty_dir:
                            print(test_dir)
                            shutil.rmtree(test_dir)
    print('All empty folders deleted successfully!')

# *************************** ECG Test Checks *************************************#
def ecg_switch_test_results(pcb_name, ecg_test_dir, test_dir_name, result_dict):
    expected_on_vpp = 2.6
    expected_off_vpp = 1e-3
    ecg_file_name = common.ecg_stream_file_name.strip('.csv') + '_SwitchOFF.csv'
    off_ecg_file_path = os.path.join(ecg_test_dir, ecg_file_name)
    ecg_file_name = common.ecg_stream_file_name.strip('.csv') + '_SwitchON.csv'
    on_ecg_file_path = os.path.join(ecg_test_dir, ecg_file_name)
    if os.path.isfile(on_ecg_file_path) and os.path.isfile(off_ecg_file_path):
        off_vpp = meas_check.calc_vpp(off_ecg_file_path)
        on_vpp = meas_check.calc_vpp(on_ecg_file_path)
        if pcb_name not in result_dict.keys():
            result_dict[pcb_name] = {}
        if 'ecg' not in result_dict[pcb_name].keys():
            result_dict[pcb_name]['ecg'] = {}
        result_dict[pcb_name]['ecg']['switch_test'] = {'meas_on_vpp': on_vpp,
                                                    'meas_off_vpp': off_vpp,
                                                    'exp_on_vpp': expected_on_vpp,
                                                    'exp_off_vpp': expected_off_vpp,
                                                    'test_dir': test_dir_name}
    return result_dict


def ecg_noise_test_results(pcb_name, ecg_test_dir, test_dir_name, result_dict):
    exp_v_noise = 50e-06
    ecg_file_name = common.ecg_stream_file_name.strip('.csv') + '_ad8233_noise.csv'
    noise_ecg_file_path = os.path.join(ecg_test_dir, ecg_file_name)
    if os.path.isfile(noise_ecg_file_path):
        v_noise = meas_check.calc_noise(noise_ecg_file_path)
        if pcb_name not in result_dict.keys():
            result_dict[pcb_name] = {}
        if 'ecg' not in result_dict[pcb_name].keys():
            result_dict[pcb_name]['ecg'] = {}
        result_dict[pcb_name]['ecg']['noise_test'] = {'exp_v_noise': exp_v_noise,
                                                      'meas_v_noise': v_noise,
                                                      'test_dir': test_dir_name}
    return result_dict


def ecg_cmrr_test_results(pcb_name, ecg_test_dir, test_dir_name, result_dict):
    exp_cmrr = '<=80'
    cmrr_a_file_name = common.ecg_stream_file_name.strip('.csv') + '_cmrr_a.csv'
    cmrr_b_file_name = common.ecg_stream_file_name.strip('.csv') + '_cmrr_b.csv'
    cmrr_c_file_name = common.ecg_stream_file_name.strip('.csv') + '_cmrr_c.csv'
    cmrr_a_file_path = os.path.join(ecg_test_dir, cmrr_a_file_name)
    cmrr_b_file_path = os.path.join(ecg_test_dir, cmrr_b_file_name)
    cmrr_c_file_path = os.path.join(ecg_test_dir, cmrr_c_file_name)
    if os.path.isfile(cmrr_a_file_path) and \
            os.path.isfile(cmrr_b_file_path) and \
            os.path.isfile(cmrr_c_file_path):
        cmrr_a = meas_check.calc_cmrr(cmrr_a_file_path)
        cmrr_b = meas_check.calc_cmrr(cmrr_b_file_path)
        cmrr_c = meas_check.calc_cmrr(cmrr_c_file_path)
        if pcb_name not in result_dict.keys():
            result_dict[pcb_name] = {}
        if 'ecg' not in result_dict[pcb_name].keys():
            result_dict[pcb_name]['ecg'] = {}
        result_dict[pcb_name]['ecg']['cmrr_test'] = {'exp_cmrr': exp_cmrr,
                                                     'cmrr_a': cmrr_a,
                                                     'cmrr_b': cmrr_b,
                                                     'cmrr_c': cmrr_c,
                                                     'test_dir': test_dir_name}
    return result_dict


def ecg_dynamic_range_test_results(pcb_name, ecg_test_dir, test_dir_name, result_dict):
    amp_vpp_list = [0.25, 0.15, 0.1, 0.05]
    freq_hz = 5
    offset_v = 1.1
    for amp_vpp in amp_vpp_list:
        ecg_file_name = common.ecg_stream_file_name.strip('.csv') + '_DR_{}Hz_{}Vpp.csv'.format(freq_hz, amp_vpp)
        ecg_file_path = os.path.join(ecg_test_dir, ecg_file_name)
        if os.path.isfile(ecg_file_path):
            gain, gain_err, gain_err_percent = meas_check.calc_dr(ecg_file_path)
            if pcb_name not in result_dict.keys():
                result_dict[pcb_name] = {}
            if 'ecg' not in result_dict[pcb_name].keys():
                result_dict[pcb_name]['ecg'] = {}
            if 'dyn_range_test' not in result_dict[pcb_name]['ecg']:
                result_dict[pcb_name]['ecg']['dyn_range_test'] = {}
            result_dict[pcb_name]['ecg']['dyn_range_test']['gain_{}'.format(amp_vpp)] = gain
            result_dict[pcb_name]['ecg']['dyn_range_test']['gain_err_{}'.format(amp_vpp)] = gain_err
            result_dict[pcb_name]['ecg']['dyn_range_test']['gain_err_per{}'.format(amp_vpp)] = gain_err_percent
            result_dict[pcb_name]['ecg']['dyn_range_test']['test_dir'] = test_dir_name
    return result_dict


def generate_results_dict(pcb_list, test_stat_dict):
    result_dict = {}
    ble_adv_battery_file_name = 'ble_advertising_curr_consumption.csv'
    shipment_battery_file_name = 'shipment_mode_curr_consumption.csv'
    streaming_battery_file_name = 'streaming_stage_curr_consumption.csv'
    for pcb in pcb_list:
        pcb_name = pcb.strip('_test_result')
        if pcb_name not in result_dict.keys():
            result_dict[pcb_name] = {}
        pcb_report_dir = os.path.join(shared_drive_dir, pcb)
        if os.path.exists(pcb_report_dir):
            print('>>>>>> {}'.format(pcb_name))
            for test_dir_name in os.listdir(pcb_report_dir):
                test_dir = os.path.join(pcb_report_dir, test_dir_name)
                if 'ecg_test' in os.listdir(test_dir):
                    if 'log.html' in os.listdir(test_dir):
                        ecg_test_dir = os.path.join(test_dir, 'ecg_test')
                        if test_stat_dict['ecg']['Switch Test']:
                            result_dict = ecg_switch_test_results(pcb_name, ecg_test_dir,
                                                                      test_dir_name, result_dict)
                        if test_stat_dict['ecg']['Noise Test']:
                            result_dict = ecg_noise_test_results(pcb_name, ecg_test_dir,
                                                                     test_dir_name, result_dict)
                        if test_stat_dict['ecg']['CMRR Test']:
                            result_dict = ecg_cmrr_test_results(pcb_name, ecg_test_dir,
                                                                    test_dir_name, result_dict)
                        if test_stat_dict['ecg']['Dynamic Range Test']:
                            result_dict = ecg_dynamic_range_test_results(pcb_name, ecg_test_dir,
                                                                             test_dir_name, result_dict)
                if 'battery_test' in os.listdir(test_dir):
                    if 'log.html' in os.listdir(test_dir):
                        battery_test_dir = os.path.join(test_dir, 'battery_test')
                        if test_stat_dict['battery']['BLE Advertising Test']:
                            result_dict = battery_test_results(pcb_name, ble_adv_battery_file_name,
                                                               battery_test_dir, test_dir_name,
                                                               result_dict, 'ble_adv_test')
                        if test_stat_dict['battery']['Shipment Mode Test']:
                            result_dict = battery_test_results(pcb_name, shipment_battery_file_name,
                                                               battery_test_dir, test_dir_name,
                                                               result_dict, 'shipment_test')
                        if test_stat_dict['battery']['Streaming Stage Test']:
                            result_dict = battery_test_results(pcb_name, streaming_battery_file_name,
                                                               battery_test_dir, test_dir_name,
                                                               result_dict, 'streaming_test')
                        # print(test_dir)
    return result_dict


# *************************** Battery Test Checks *************************************#
def battery_test_results(pcb_name, battery_file_name, battery_test_dir,
                          test_dir_name, result_dict, result_key):
    exp_curr = 17e-3
    battery_file_path = os.path.join(battery_test_dir, battery_file_name)
    if os.path.isfile(battery_file_path):
        check_pass, avg_curr, failure_curr = meas_check.check_battery_charge(battery_file_path, curr_threshold=exp_curr)
        if pcb_name not in result_dict.keys():
            result_dict[pcb_name] = {}
        if 'battery' not in result_dict[pcb_name].keys():
            result_dict[pcb_name]['battery'] = {}
        result_dict[pcb_name]['battery'][result_key] = {'check_pass': check_pass,
                                                        'avg_curr': avg_curr,
                                                        'exp_curr': '<{:.4f}'.format(exp_curr),
                                                        'failure_curr': failure_curr,
                                                        'test_dir': test_dir_name}
    return result_dict


def generate_battery_results(pcb_list, battery_test_stat_dict):
    battery_result_dict = {}
    ble_adv_battery_file_name = 'ble_advertising_curr_consumption.csv'
    shipment_battery_file_name = 'shipment_mode_curr_consumption.csv'
    streaming_battery_file_name = 'streaming_stage_curr_consumption.csv'
    for pcb in pcb_list:
        pcb_name = pcb.strip('_test_result')
        pcb_report_dir = os.path.join(shared_drive_dir, pcb)
        if os.path.exists(pcb_report_dir):
            print('>>>>>> {}'.format(pcb_name))
            for test_dir_name in os.listdir(pcb_report_dir):
                test_dir = os.path.join(pcb_report_dir, test_dir_name)
                if 'battery_test' in os.listdir(test_dir):
                    if 'log.html' in os.listdir(test_dir):
                        battery_test_dir = os.path.join(test_dir, 'battery_test')
                        if battery_test_stat_dict['BLE Advertising Test']:
                            battery_result_dict = battery_test_results(pcb_name, ble_adv_battery_file_name,
                                                                       battery_test_dir, test_dir_name,
                                                                       battery_result_dict, 'ble_adv_test')
                        if battery_test_stat_dict['Shipment Mode Test']:
                            battery_result_dict = battery_test_results(pcb_name, shipment_battery_file_name,
                                                                       battery_test_dir, test_dir_name,
                                                                       battery_result_dict, 'shipment_test')
                        if battery_test_stat_dict['Streaming Stage Test']:
                            battery_result_dict = battery_test_results(pcb_name, streaming_battery_file_name,
                                                                       battery_test_dir, test_dir_name,
                                                                       battery_result_dict, 'streaming_test')
                        # print(test_dir)
    print('Battery Test Results:', battery_result_dict.keys())
    return battery_result_dict


def generate_report(consolidated_result_dict):
    curr_dir = os.getcwd()
    template_path = os.path.join(curr_dir, '../report_template.xlsx')
    report_path = os.path.join(curr_dir, '../report.xlsx')
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    wb = load_workbook(template_path)
    ws = wb.get_active_sheet()
    exp_val_col_idx = 4
    col_idx = 5
    for pcb_name, pcb_result_dict in consolidated_result_dict.iteritems():  # ECG Test Results Parse Loop
        row_idx = 12
        ws.cell(row=row_idx, column=col_idx).value = pcb_name
        ws.cell(row=row_idx, column=col_idx).border = thin_border
        ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color="85A3E0", end_color="85A3E0", fill_type="solid")
        if pcb_result_dict:
            if 'ecg' in pcb_result_dict:
                result_dict = pcb_result_dict['ecg']
                if 'switch_test' in result_dict.keys():
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['switch_test']['meas_on_vpp']
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = result_dict['switch_test']['exp_on_vpp']
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['switch_test']['meas_off_vpp']
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = result_dict['switch_test']['exp_off_vpp']
                else:
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                if 'noise_test' in result_dict.keys():
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['noise_test']['meas_v_noise']
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = result_dict['noise_test']['exp_v_noise']
                else:
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                if 'cmrr_test' in result_dict.keys():
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = result_dict['cmrr_test']['exp_cmrr']
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['cmrr_test']['cmrr_a']
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = result_dict['cmrr_test']['exp_cmrr']
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['cmrr_test']['cmrr_b']
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = result_dict['cmrr_test']['exp_cmrr']
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['cmrr_test']['cmrr_c']
                else:
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                if 'dyn_range_test' in result_dict.keys():
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = NA
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['dyn_range_test']['gain_0.25']
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = NA
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['dyn_range_test']['gain_err_0.25']
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = '<10'
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['dyn_range_test']['gain_err_per0.25']
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = NA
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['dyn_range_test']['gain_0.15']
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = NA
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['dyn_range_test']['gain_err_0.15']
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = '<10'
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['dyn_range_test']['gain_err_per0.15']
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = NA
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['dyn_range_test']['gain_0.1']
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = NA
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['dyn_range_test']['gain_err_0.1']
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = '<10'
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['dyn_range_test']['gain_err_per0.1']
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = NA
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['dyn_range_test']['gain_0.05']
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = NA
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['dyn_range_test']['gain_err_0.05']
                    row_idx += 1
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = '<10'
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['dyn_range_test']['gain_err_per0.05']
                else:
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
            else:
                total_ecg_rows = 18
                for i in range(total_ecg_rows):
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
            if 'battery' in pcb_result_dict:
                result_dict = pcb_result_dict['battery']
                if 'ble_adv_test' in result_dict.keys():
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['ble_adv_test']['avg_curr']
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = result_dict['ble_adv_test']['exp_curr']
                else:
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                if 'shipment_test' in result_dict.keys():
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['shipment_test']['avg_curr']
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = result_dict['shipment_test']['exp_curr']
                else:
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
                if 'streaming_test' in result_dict.keys():
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = result_dict['streaming_test']['avg_curr']
                    ws.cell(row=row_idx, column=exp_val_col_idx).value = result_dict['streaming_test']['exp_curr']
                else:
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
            else:
                total_battery_rows = 3
                for i in range(total_battery_rows):
                    row_idx += 1
                    ws.cell(row=row_idx, column=col_idx).value = NA
        else:
            total_rows = 21
            for i in range(total_rows):
                row_idx += 1
                ws.cell(row=row_idx, column=col_idx).value = NA
        col_idx += 1
    wb.save(report_path)
    return report_path


def initialize_report_gen():
    test_stat_dict = {'ecg': {'Switch Test': False, 'Noise Test': False,
                              'CMRR Test': False, 'Dynamic Range Test': False},
                      'battery': {'BLE Advertising Test': False, 'Shipment Mode Test': False,
                                  'Streaming Stage Test': False}}
    choice_list = easygui.multchoicebox('Welcome to PCB Test Report Generation wizard!\n\n'
                                        'Select all the required test cases from the below list for '
                                        'generating report.',
                                        'PCB Test Report Generator',
                                        ['ECG Test - Switch Test',
                                         'ECG Test - Noise Test',
                                         'ECG Test - CMRR Test',
                                         'ECG Test - Dynamic Range Test',
                                         'Battery Test - BLE Advertising Test',
                                         'Battery Test - Shipment Stage Test',
                                         'Battery Test - Streaming Stage Test'],
                                        0)
    if choice_list:
        for choice in choice_list:
            if 'ECG Test - ' in choice:
                choice = choice.split(' - ')[-1]
                test_stat_dict['ecg'][choice] = True
            elif 'Battery Test - ' in choice:
                choice = choice.split(' - ')[-1]
                test_stat_dict['battery'][choice] = True
    else:
        test_stat_dict = None
    return test_stat_dict


if __name__ == '__main__':
    pcb_folder_list = ['0E630A3D5CC6_test_result', '3DFA8E4D22F4_test_result', '6ED7EE20EFED_test_result',
                '15D955EF64D5_test_result', '3971C71D62F0_test_result', '839510F51EFD_test_result']
    # delete_empty_report_folders(shared_drive_dir)
    test_stat_dict = initialize_report_gen()
    if test_stat_dict:
        # ecg_result_dict = generate_ecg_results(pcb_folder_list, test_stat_dict['ecg'])
        # battery_result_dict = generate_battery_results(pcb_folder_list, test_stat_dict['battery'])
        result_dict = generate_results_dict(pcb_folder_list, test_stat_dict)
        report_path = generate_report(result_dict)
        print(result_dict)
        # report_path = generate_report(ecg_result_dict)
        print('Report generation completed!')
        print('Report Path: ' + os.path.abspath(report_path))
    else:
        print('Report generation completed!')
        print('No report generated due to 0 test case selection!')
